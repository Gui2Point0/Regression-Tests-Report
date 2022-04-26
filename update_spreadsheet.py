from openpyxl import load_workbook

filename = 'RegressionTestReports'

def CreateNewSheet(sheetName, wb):
    sheetsCount = len(wb.sheetnames)
    wb.create_sheet(sheetName, sheetsCount+1)
    ws = wb[sheetName]

    return ws

def OpenExcelFile(sheetName, resultList):
    #Loading existing Excel File into work_book Object
    wb = load_workbook(filename + '.xlsx') 
    ws = wb.active

    #Creating a new Sheet at next position (index) available
    ws = CreateNewSheet(sheetName, wb)

    #Add first row as columns info
    ws.append(['Test Name', 'Score', 'Duration'])

    #Add the results to the sheet
    for row in resultList:
        ws.append(row)

    #Saving Excel File
    wb.save(filename + '.xlsx')

