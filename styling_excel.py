from itertools import count
from numpy import empty
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import load_workbook
from sqlalchemy import null, true

#Make titles font bold
def MakeTextBold(cell,ws):
    ws[cell].font = Font(bold=True)

#Align titles to the center
def CenterText(cell, ws):
    ws[cell].alignment = Alignment(horizontal='center', vertical='center')

#Pass and Fail background
def CellBackgroung(ws): 
    green = '90EE90' 
    red = 'FF6961'
    counter = 1

    while(True):
        counter += 1
        rslt_cell = ws['B' + str(counter)]
    
        if rslt_cell.value != None:
            if rslt_cell.value == 'pass':
                rslt_cell.fill = PatternFill('solid', fgColor=green)
            elif rslt_cell.value == 'fail':
                rslt_cell.fill = PatternFill('solid', fgColor=red)
        else:
            break


#Styling columns info title
def ColumnTitles(ws):
    CenterText('A1', ws)
    MakeTextBold('A1', ws)
    
    CenterText('B1', ws)
    MakeTextBold('B1', ws)
    
    CenterText('C1', ws)
    MakeTextBold('C1', ws)

##########################################################################################################################

#Main
def SheetStyles(ws):

    #Call methods
    ColumnTitles(ws)
    CellBackgroung(ws)
